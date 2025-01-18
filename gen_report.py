import pandas as pd

from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

import pdb


def new_primitive_count():
    """
    Return a new dict for each supported data type used in ONNX. This is for tracking all compute primitive in its respective data type

    Returns:
        dict: dict of supported data type
    """
    return {
        "FLOAT": 0,
        "UINT8": 0,
        "INT8": 0,
    }


class ReportGenerator:
    """
    This class is taking the model data from ModelStats and perform additional model analysis

    Attributes:
    model_stats (list):        The basic statistics of ONNX model
    xlsx_filename (str):        report filename
    """

    def __init__(self, model_stats, xlsx_filename):
        self.model_stats = model_stats
        self.xlsx_filename = xlsx_filename

    def write_xlsx(self):
        with pd.ExcelWriter(self.xlsx_filename) as writer:
            model_frame = pd.DataFrame(self.model_stats)
            supported_model_frame = model_frame[
                model_frame["Supported"] != False
            ].copy()

            # Remove "Supported" from the original columns
            supported_model_frame.drop(columns=["Supported"], inplace=True)
            stat_names = [
                key for key in self.model_stats[0].keys() if key != "Supported"
            ]

            TOTAL_STAT_KEYS = {
                "MAC Count",
                "ALU Count",
                "EXP Count",
                "LOG Count",
                "DIV Count",
                "Input Size (bytes)",
                "Weight Size (bytes)",
                "Output Size (bytes)",
            }

            totals = {key: 0 for key in stat_names if key in TOTAL_STAT_KEYS}

            for stat in self.model_stats:
                for key in totals:
                    if key in stat:
                        totals[key] += stat[key]

            totals = {key: (totals[key] if key in totals else "") for key in stat_names}
            totals.update({"Operator Name": "Total"})
            supported_model_frame = pd.concat(
                [supported_model_frame, pd.DataFrame([totals])], ignore_index=True
            )

            supported_model_frame.index.name = "Index"

            supported_model_frame.to_excel(writer, sheet_name="ONNX Model Breakdown")

        workbook = load_workbook(self.xlsx_filename)
        sheet = workbook["ONNX Model Breakdown"]
        for column in sheet.columns:
            col_letter = column[0].column_letter
            sheet.column_dimensions[col_letter].width = 16

        if len(supported_model_frame.columns) > 26:
            table_model_stats_range = f"A1:{chr(64 + len(supported_model_frame.columns) // 26) + chr(65 + len(supported_model_frame.columns) % 26 - 1)}{supported_model_frame.shape[0] + 1}"
        else:
            table_model_stats_range = f"A1:{chr(65 + len(supported_model_frame.columns))}{supported_model_frame.shape[0] + 1}"

        table_ops_list = Table(displayName="ONNX_Ops_List", ref=table_model_stats_range)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table_ops_list.tableStyleInfo = style
        sheet.add_table(table_ops_list)

        # Adding a new new sheet to summarize op_type
        model_sheet_data = pd.DataFrame(self.model_stats)

        for col in [
            "MAC Count",
            "ALU Count",
            "EXP Count",
            "LOG Count",
            "DIV Count",
            "Input Size (bytes)",
            "Weight Size (bytes)",
            "Output Size (bytes)",
        ]:
            model_sheet_data[col] = pd.to_numeric(
                model_sheet_data[col], errors="coerce"
            )

        model_sheet_data["Supported"] = model_sheet_data["Supported"].apply(
            lambda x: "TRUE" if x == 1 else "FALSE"
        )

        ops_summary_frame = (
            model_sheet_data.groupby("Op Type")
            .agg(
                {
                    "Supported": "first",
                    "MAC Count": "sum",
                    "ALU Count": "sum",
                    "EXP Count": "sum",
                    "LOG Count": "sum",
                    "DIV Count": "sum",
                    "Input Size (bytes)": "sum",
                    "Weight Size (bytes)": "sum",
                    "Output Size (bytes)": "sum",
                }
            )
            .reset_index()
        )

        ops_summary_frame["Operator Count"] = (
            model_sheet_data["Op Type"]
            .value_counts()
            #            .reindex(ops_summary_frame["Op Type"])
            .values
        )

        columns = ["Op Type", "Operator Count"] + [
            col
            for col in ops_summary_frame.columns
            if col not in ["Op Type", "Operator Count"]
        ]
        ops_summary_frame = ops_summary_frame[columns]

        ops_summary_frame.loc["Total"] = ops_summary_frame.sum(numeric_only=True)
        ops_summary_frame.loc["Total", "Op Type"] = "Total"

        summary_rows = [
            ops_summary_frame.columns.tolist()
        ] + ops_summary_frame.values.tolist()

        ops_sheet = workbook.create_sheet(title="Op Type Summary")
        for row in summary_rows:
            ops_sheet.append(row)

        for column in ops_sheet.columns:
            col_letter = column[0].column_letter
            ops_sheet.column_dimensions[col_letter].width = 16

        table_model_stats_range = f"A1:{chr(64 + len(ops_summary_frame.columns))}{ops_summary_frame.shape[0] + 1}"

        ops_summary_list = Table(
            displayName="Ops_Summary_List", ref=table_model_stats_range
        )
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ops_summary_list.tableStyleInfo = style
        ops_sheet.add_table(ops_summary_list)

        workbook.save(self.xlsx_filename)
